import os
import pymysql
from openpyxl import Workbook
from datetime import datetime
from sql_file_list_MS import file_list
from db_config import LSH_CONFIG, MFS_CONFIG, load_environment
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

def sanitize_cell(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def sanitize_excel_value(val, max_length=300):
    if isinstance(val, str):
        # Remove characters that are illegal in XML (which Excel uses)
        val = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', val)
        
        # Remove non-BMP Unicode characters (e.g., certain emoji)
        val = ''.join(c for c in val if ord(c) <= 0xFFFF)
        
        # Truncate if it's too long
        if len(val) > max_length:
            val = val[:max_length] + '... [truncated]'
    return val

start_time = datetime.now()

CONFIG = LSH_CONFIG
env = load_environment(CONFIG)
source_connection = env["source_connection"]
archived_connection = env["archived_connection"]
db_placeholders = env["placeholders"]

wb = Workbook(write_only=True)
ws_summary = wb.create_sheet(title="Verification Results")
ws_summary.append([
    "Overall Result", "File", "Table/s", "Residual Count", "Residual Status",
    "Retained Count", "Expected Count", "Actual vs Expected Retained Result",
    "Data Integrity", "Timestamp"
])

queries_dir = "queries_ms"
results_dir = "results_ms"
sql_files = []
missing_files = []

for f in file_list:
    full_path = os.path.join(queries_dir, f)
    if not os.path.isfile(full_path):
        missing_files.append(f)
    else:
        sql_files.append(full_path)

if missing_files:
    print("\n❌ ERROR: The following files were not found in the 'queries_ms' folder:")
    for mf in missing_files:
        print(f"   - {mf}")
    raise FileNotFoundError("One or more SQL files are missing. Please check the 'queries_ms' folder.")

# ✅ Totals initialization
total_rows_verified = 0
total_columns_verified = 0

for sql_file_path in sql_files:
    sql_file_name = os.path.basename(sql_file_path)

    try:
        with open(sql_file_path, 'r') as file:
            lines = file.readlines()
    except Exception as e:
        print(f"[X] Could not read file {sql_file_path}: {e}")
        continue

    print(f"\n📄 Verifying file: {sql_file_path}")
    
    # Mapping for each (file, table)
    results_map = {}
    queries_data = {}

    i = 0
    while i < len(lines):
        if lines[i].strip().startswith("--"):
            label_line = lines[i].strip()[2:].strip()
            label = label_line.split('|')[0].strip().lower()

            i += 1
            table_names = ""
            if i < len(lines) and lines[i].strip().startswith("--"):
                table_names = lines[i].strip()[2:].strip()
                i += 1

            sql_lines = []
            while i < len(lines) and not lines[i].strip().startswith("--"):
                sql_lines.append(lines[i])
                i += 1

            sql_query = ''.join(sql_lines).strip()

            for placeholder, value in db_placeholders.items():
                sql_query = sql_query.replace(f"{{{placeholder}}}", value)

            if not sql_query:
                continue

            conn = archived_connection if 'retained' in label else source_connection

            try:
                with conn.cursor() as cursor:
                    cursor.execute(sql_query)
                    rows = cursor.fetchall()
                    count = len(rows)
                    print(f"[✓] {label.upper()} → {count}")

                    key = (sql_file_name, table_names)
                    if key not in results_map:
                        results_map[key] = {
                            "residual_check": None,
                            "retained_count": None,
                            "expected_count": None,
                            "retained_rows": [],
                            "expected_rows": [],
                            "retained_columns": [],
                            "expected_columns": []
                        }

                    if 'residual' in label:
                        results_map[key]["residual_check"] = count
                    elif 'retained' in label:
                        results_map[key]["retained_count"] = count
                        results_map[key]["retained_rows"] = list(rows)
                        results_map[key]["retained_columns"] = [desc[0] for desc in cursor.description]
                    elif 'expected' in label:
                        results_map[key]["expected_count"] = count
                        results_map[key]["expected_rows"] = list(rows)
                        results_map[key]["expected_columns"] = [desc[0] for desc in cursor.description]

            except Exception as e:
                print(f"[X] Error in {label.upper()} → {e}")
                ws_summary.append([
                    "FAIL", sql_file_name, table_names,
                    "Error" if 'residual' in label else "",
                    "FAIL",
                    "Error" if 'retained' in label else "",
                    "Error" if 'expected' in label else "",
                    "FAIL", f"Error: {str(e)}",
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
        else:
            i += 1

    for (sql_file, table), data in results_map.items():
        residual = data["residual_check"]
        retained = data["retained_count"]
        expected = data["expected_count"]
        retained_rows = data["retained_rows"]
        expected_rows = data["expected_rows"]
        column_names_retained = [c + "_retained" for c in data["retained_columns"]]
        column_names_expected = [c + "_expected" for c in data["expected_columns"]]

        safe_sheet_name = re.sub(r'[\\/*?:[\]]', "", sql_file[:25])  # Excel sheet naming limit
        if len(wb.sheetnames) >= 1:
            safe_sheet_name += f"_{len(wb.sheetnames)}"

        ws_data = wb.create_sheet(title=safe_sheet_name)

        mismatch_count = 0

        interleaved_column_names = ["Overall Status"]
        for r_col, e_col in zip(column_names_retained, column_names_expected):
            col_base = r_col.replace("_retained", "")
            interleaved_column_names.extend([f"{col_base}_status", r_col, e_col])
        ws_data.append([sanitize_cell(c) for c in interleaved_column_names])

        base_col_names = [col.replace("_retained", "") for col in column_names_retained]
        key_indexes = [0] + [i for i, name in enumerate(base_col_names) if 'id' in name.lower()]

        def row_to_key(row, indexes):
            return tuple(row[i] for i in indexes)

        retained_dict = {row_to_key(row, key_indexes): row for row in retained_rows}
        expected_dict = {row_to_key(row, key_indexes): row for row in expected_rows}
        all_keys = set(retained_dict.keys()).union(expected_dict.keys())

        for key in all_keys:
            retained_row = retained_dict.get(key, [""] * len(column_names_retained))
            expected_row = expected_dict.get(key, [""] * len(column_names_expected))

            is_retained_missing = all(v == "" for v in retained_row)
            is_expected_missing = all(v == "" for v in expected_row)

            interleaved_row = ["MISSING"] if is_retained_missing or is_expected_missing else ["PASS"]

            for r_val, e_val in zip(retained_row, expected_row):
                r_str = str(r_val).strip().lower() if isinstance(r_val, str) else str(r_val).strip()
                e_str = str(e_val).strip().lower() if isinstance(e_val, str) else str(e_val).strip()

                status = "matched" if r_str == e_str else "mismatched"
                if status == "mismatched" and interleaved_row[0] == "PASS":
                    interleaved_row[0] = "FAIL"
                    mismatch_count += 1
                interleaved_row.extend([status, r_val, e_val])

            total_rows_verified += 1
            try:
                ws_data.append([sanitize_excel_value(c) for c in interleaved_row])
            except IllegalCharacterError as e:
                print(f"⚠️ Illegal character in Excel row for file {sql_file_name}: {e}")

        if retained and expected:
            total_columns_verified += len(base_col_names)

        data_integrity_display = (
            "No Data Fetched" if (retained == expected == None or (retained == 0 and expected == 0)) else
            "PASS" if mismatch_count == 0 and retained == expected else
            f"{mismatch_count} mismatches found" if mismatch_count else
            f"{abs(retained - expected)} data affected"
        )

        retained_expected_status = "PASS" if mismatch_count == 0 and retained == expected else "FAIL"
        residual_status = "PASS" if residual == 0 else "FAIL"

        overall_result = "PASS" if (
            residual_status == "PASS" and
            mismatch_count == 0 and
            retained == expected
        ) else "FAIL"

        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary.append([
            overall_result, sql_file, table,
            residual if residual is not None else "",
            residual_status,
            retained if retained is not None else "",
            expected if expected is not None else "",
            retained_expected_status,
            data_integrity_display,
            timestamp_str
        ])

# Final output save
os.makedirs(results_dir, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
CONFIG_NAME = "LSH" if CONFIG == LSH_CONFIG else "MFS"
output_file = os.path.join(results_dir, f"{CONFIG_NAME}_query_results_{timestamp}.xlsx")
wb.save(output_file)

duration = datetime.now() - start_time
total_seconds = duration.total_seconds()
milliseconds = int((total_seconds - int(total_seconds)) * 1000)
formatted_duration = f"{str(duration).split('.')[0]}.{milliseconds:03d}"

print(f"\n📉 Results saved to: {output_file}")
print(f"⏱️ Duration: {formatted_duration}")
print(f"🔢 Total rows verified: {total_rows_verified}")
print(f"📊 Total columns verified: {total_columns_verified}")
