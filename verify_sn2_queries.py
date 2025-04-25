import os
import pymysql
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from datetime import datetime
from db_config import load_environment, LSH_CONFIG  # or MFS_CONFIG
from sql_file_list_SN2 import file_list

# Configuration
BU = 'sn2_lsh'
queries_dir = "queries_sn2"
results_dir = "results_sn2"

start_time = datetime.now()

env = load_environment(LSH_CONFIG)  # change to MFS_CONFIG if needed
source_connection = env["source_connection"]
archived_connection = env["archived_connection"]
placeholders = env["placeholders"]

wb = Workbook(write_only=True)
ws_summary = wb.create_sheet(title="Verification Results")
ws_summary.append([
    "Overall Result",
    "File",
    "Table/s",
    "Residual Count",
    "Residual Status",
    "Retained Count",
    "Expected Count",
    "Actual vs Expected Retained Result",
    "Data Integrity",
    "Timestamp"
])

sql_files = []
missing_files = []

for f in file_list:
    full_path = os.path.join(queries_dir, f)
    if not os.path.isfile(full_path):
        missing_files.append(f)
    else:
        sql_files.append(full_path)

if missing_files:
    print("\n❌ ERROR: The following files were not found in the 'queries' folder:")
    for mf in missing_files:
        print(f"   - {mf}")
    raise FileNotFoundError("One or more SQL files are missing. Please check the 'queries' folder.")

# ✅ Totals initialization
total_rows_verified = 0
total_columns_verified = 0

for sql_file_path in sql_files:
    sql_file_name = os.path.basename(sql_file_path)

    try:
        with open(sql_file_path, 'r') as file:
            lines = file.readlines()
    except Exception as e:
        print(f"[X] Could not read file {sql_file_path}: {e}")
        continue

    print(f"\n📄 Verifying file: {sql_file_path}")

    ws_data = wb.create_sheet(title=sql_file_name[:31])

    results_map = {}
    queries_data = {}
    column_names_retained = []
    column_names_expected = []
    i = 0

    while i < len(lines):
        if lines[i].strip().startswith("--"):
            label_line = lines[i].strip()[2:].strip()
            label = label_line.split('|')[0].strip().lower()

            i += 1
            table_names = ""
            if i < len(lines) and lines[i].strip().startswith("--"):
                table_names = lines[i].strip()[2:].strip()
                i += 1

            sql_lines = []
            while i < len(lines) and not lines[i].strip().startswith("--"):
                sql_lines.append(lines[i])
                i += 1
            sql_query = ''.join(sql_lines).strip()

            # Replace all placeholders dynamically
            for placeholder, value in placeholders.items():
                sql_query = sql_query.replace(f"{{{placeholder}}}", value)

            if not sql_query:
                continue

            conn = archived_connection if 'retained' in label else source_connection

            try:
                with conn.cursor() as cursor:
                    cursor.execute(sql_query)
                    rows = cursor.fetchall()
                    count = len(rows)
                    print(f"[✓] {label.upper()} → {count}")

                    if (sql_file_name, table_names) not in results_map:
                        results_map[(sql_file_name, table_names)] = {
                            "residual_check": None,
                            "retained_count": None,
                            "expected_count": None
                        }

                    if 'residual' in label:
                        results_map[(sql_file_name, table_names)]["residual_check"] = count
                    elif 'retained' in label:
                        results_map[(sql_file_name, table_names)]["retained_count"] = count
                        queries_data['retained'] = list(rows)
                        column_names_retained = [desc[0] + "_retained" for desc in cursor.description]
                    elif 'expected' in label:
                        results_map[(sql_file_name, table_names)]["expected_count"] = count
                        queries_data['expected'] = list(rows)
                        column_names_expected = [desc[0] + "_expected" for desc in cursor.description]

            except Exception as e:
                print(f"[X] Error in {label.upper()} → {e}")
                ws_summary.append([
                    "FAIL",
                    sql_file_name,
                    table_names,
                    "Error" if 'residual' in label else "",
                    "FAIL",
                    "Error" if 'retained' in label else "",
                    "Error" if 'expected' in label else "",
                    "FAIL",
                    "Error: " + str(e),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
        else:
            i += 1

    for (sql_file, table), counts in results_map.items():
        residual = counts["residual_check"]
        retained = counts["retained_count"]
        expected = counts["expected_count"]

        residual_status = "PASS" if residual == 0 else "FAIL"

        retained_rows = queries_data.get("retained", [])
        expected_rows = queries_data.get("expected", [])

        mismatch_count = 0

        interleaved_column_names = ["Overall Status"]
        for r_col, e_col in zip(column_names_retained, column_names_expected):
            col_base = r_col.replace("_retained", "")
            interleaved_column_names.append(f"{col_base}_status")
            interleaved_column_names.append(r_col)
            interleaved_column_names.append(e_col)
        ws_data.append(interleaved_column_names)

        base_col_names = [col.replace("_retained", "") for col in column_names_retained]
        key_indexes = [0] + [i for i, name in enumerate(base_col_names) if 'id' in name.lower()]

        def row_to_key(row, indexes):
            return tuple(row[i] for i in indexes)

        retained_dict = {row_to_key(row, key_indexes): row for row in retained_rows}
        expected_dict = {row_to_key(row, key_indexes): row for row in expected_rows}

        all_keys = set(retained_dict.keys()).union(expected_dict.keys())

        for key in all_keys:
            retained_row = retained_dict.get(key, [""] * len(column_names_retained))
            expected_row = expected_dict.get(key, [""] * len(column_names_expected))

            is_retained_missing = all(v == "" for v in retained_row)
            is_expected_missing = all(v == "" for v in expected_row)

            if is_retained_missing or is_expected_missing:
                interleaved_row = ["MISSING"]
            else:
                interleaved_row = ["PASS"]

            for r_val, e_val in zip(retained_row, expected_row):
                status = "matched" if r_val == e_val else "mismatched"
                if status == "mismatched" and interleaved_row[0] == "PASS":
                    interleaved_row[0] = "FAIL"
                    mismatch_count += 1
                interleaved_row.append(status)
                interleaved_row.append(r_val)
                interleaved_row.append(e_val)

            total_rows_verified += 1
            ws_data.append(interleaved_row)

        if (retained and expected):
            total_columns_verified += len(base_col_names)

        if (retained is None or retained == 0) and (expected is None or expected == 0):
            data_integrity_display = "No Data Fetched"
        elif retained != expected:
            data_integrity_display = f"{abs(retained - expected)} data affected"
        elif mismatch_count == 0:
            data_integrity_display = "PASS"
        else:
            data_integrity_display = f"{mismatch_count} mismatches found"

        retained_expected_status = "PASS" if mismatch_count == 0 and retained == expected else "FAIL"

        if data_integrity_display == "No Data Fetched" and residual_status == "PASS" and retained_expected_status == "PASS":
            overall_result = "PASS"
        else:
            overall_result = "FAIL" if any([
                residual_status != "PASS",
                retained_expected_status != "PASS",
                data_integrity_display != "PASS"
            ]) else "PASS"

        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws_summary.append([
            overall_result,
            sql_file,
            table,
            residual if residual is not None else "",
            residual_status,
            retained if retained is not None else "",
            expected if expected is not None else "",
            retained_expected_status,
            data_integrity_display,
            timestamp_str
        ])

# Save workbook
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
os.makedirs(results_dir, exist_ok=True)
output_file = os.path.join(results_dir, f"{BU}_query_results_{timestamp}.xlsx")
wb.save(output_file)

# Summary
duration = datetime.now() - start_time
total_seconds = duration.total_seconds()
milliseconds = int((total_seconds - int(total_seconds)) * 1000)
formatted_duration = f"{str(duration).split('.')[0]}.{milliseconds:03d}"

print(f"\n📉 Results saved to: {output_file}")
print(f"⏱️ Duration: {formatted_duration}")
print(f"🔢 Total rows verified: {total_rows_verified}")
print(f"📊 Total columns verified: {total_columns_verified}")