import os
import pymysql
from openpyxl import Workbook
from datetime import datetime
from db_config import get_source_connection, get_archived_connection
from sql_file_list_SN2 import file_list

BU = 'lsh'
start_time = datetime.now()

source_connection = get_source_connection(BU)
archived_connection = get_archived_connection(BU)

wb = Workbook(write_only=True)
ws_summary = wb.create_sheet(title="Verification Results")
ws_summary.append([
    "Overall Result", "File", "Table/s", "Residual Count", "Residual Status",
    "Retained Count", "Expected Count", "Actual vs Expected Retained Result",
    "Data Integrity", "Timestamp"
])

queries_dir = "queries"
sql_files = [os.path.join(queries_dir, f) for f in file_list if os.path.isfile(os.path.join(queries_dir, f))]
missing_files = [f for f in file_list if not os.path.isfile(os.path.join(queries_dir, f))]

if missing_files:
    print("\n❌ ERROR: Missing SQL files:")
    for mf in missing_files:
        print(f"   - {mf}")
    raise FileNotFoundError("Some SQL files are missing in the 'queries' folder.")

rows_verified, columns_verified = 0, 0
CHUNK_SIZE = 10000

def get_query_results(cursor, sql_query):
    cursor.execute(sql_query)
    while True:
        chunk = cursor.fetchmany(CHUNK_SIZE)
        if not chunk:
            break
        for row in chunk:
            yield row

for sql_path in sql_files:
    sql_name = os.path.basename(sql_path)
    print(f"\n📄 Verifying file: {sql_path}")
    ws_data = wb.create_sheet(title=sql_name[:31])

    try:
        with open(sql_path, 'r') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"[X] Error reading {sql_path}: {e}")
        continue

    results_map, queries_data = {}, {}
    col_retained, col_expected = [], []
    i = 0

    while i < len(lines):
        if lines[i].strip().startswith("--"):
            label = lines[i].strip()[2:].split('|')[0].strip().lower()
            i += 1
            table_names = lines[i].strip()[2:].strip() if i < len(lines) and lines[i].strip().startswith("--") else ""
            i += 1 if table_names else 0

            sql_query = ''
            while i < len(lines) and not lines[i].strip().startswith("--"):
                sql_query += lines[i]
                i += 1

            sql_query = sql_query.replace("{SOURCE_DB}", source_connection.db.decode())
            sql_query = sql_query.replace("{ARCHIVED_DB}", archived_connection.db.decode())
            if not sql_query.strip():
                continue

            conn = archived_connection if 'retained' in label else source_connection
            try:
                with conn.cursor() as cursor:
                    count = 0
                    data = {}
                    columns = []
                    for row in get_query_results(cursor, sql_query):
                        key = tuple(row)  # ✅ use full row as key to avoid overwrite
                        data[key] = row
                        count += 1
                        if not columns:
                            columns = [desc[0] + f"_{label}" for desc in cursor.description]
                    print(f"[✓] {label.upper()} → {count}")

                    results_map.setdefault((sql_name, table_names), {
                        "residual_check": None, "retained_count": None, "expected_count": None
                    })

                    if 'residual' in label:
                        results_map[(sql_name, table_names)]["residual_check"] = count
                    elif 'retained' in label:
                        results_map[(sql_name, table_names)]["retained_count"] = count
                        queries_data['retained'] = data
                        col_retained = columns
                    elif 'expected' in label:
                        results_map[(sql_name, table_names)]["expected_count"] = count
                        queries_data['expected'] = data
                        col_expected = columns
            except Exception as e:
                print(f"[X] Error in {label.upper()} → {e}")
                ws_summary.append(["FAIL", sql_name, table_names, "Error", "FAIL", "Error", "Error", "FAIL", f"Error: {e}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    for (sql_name, table), counts in results_map.items():
        residual, retained, expected = counts.values()
        residual_status = "PASS" if residual == 0 else "FAIL"

        ret_dict = queries_data.get("retained", {})
        exp_dict = queries_data.get("expected", {})
        all_keys = set(ret_dict.keys()) | set(exp_dict.keys())
        mismatch_count = 0

        header = ["Overall Status"] + [f"{col.replace('_retained', '')}_status" for col in col_retained] + col_retained + col_expected
        ws_data.append(header)

        base_cols = [col.replace("_retained", "") for col in col_retained]

        for key in all_keys:
            r_row = ret_dict.get(key, [""] * len(col_retained))
            e_row = exp_dict.get(key, [""] * len(col_expected))

            is_missing = all(v == "" for v in r_row) or all(v == "" for v in e_row)
            row_status = "MISSING" if is_missing else "PASS"
            row_data = [row_status]

            for rv, ev in zip(r_row, e_row):
                status = "matched" if rv == ev else "mismatched"
                if status == "mismatched" and row_status == "PASS":
                    row_data[0] = "FAIL"
                    mismatch_count += 1
                row_data.extend([status, rv, ev])

            rows_verified += 1
            ws_data.append(row_data)

        if retained and expected:
            columns_verified += len(base_cols)

        data_integrity = (
            "No Data Fetched" if not retained and not expected else
            f"{abs(retained - expected)} data affected" if retained != expected else
            ("PASS" if mismatch_count == 0 else f"{mismatch_count} mismatches found")
        )

        ret_exp_status = "PASS" if retained == expected and mismatch_count == 0 else "FAIL"
        overall = "PASS" if data_integrity == "No Data Fetched" and residual_status == "PASS" and ret_exp_status == "PASS" else "FAIL"

        ws_summary.append([
            overall, sql_name, table,
            residual if residual is not None else "",  # ✅ fixed Residual Count display
            residual_status, retained or "", expected or "", ret_exp_status,
            data_integrity, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

os.makedirs("results", exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = f"results/{BU}_query_results_{timestamp}.xlsx"
wb.save(output_file)

duration = datetime.now() - start_time
total_seconds = duration.total_seconds()
milliseconds = int((total_seconds - int(total_seconds)) * 1000)
print(f"\n📉 Results saved to: {output_file}")
print(f"⏱️ Duration: {str(duration).split('.')[0]}.{milliseconds:03d}")
print(f"🔢 Total rows verified: {rows_verified}")
print(f"📊 Total columns verified: {columns_verified}")
