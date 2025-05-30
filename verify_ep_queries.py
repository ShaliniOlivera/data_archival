import os
import pymysql
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from datetime import datetime
from db_config import load_environment, MFS_CONFIG, LSH_CONFIG, EP_CONFIG
from sql_file_list_EP import file_list
import re
from openpyxl.utils.exceptions import IllegalCharacterError

# Remove all non-printable characters except common whitespace
def clean_excel_value(value):
    if isinstance(value, str):
        # Allow tab (\t), newline (\n), carriage return (\r)
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return value

BU = 'EP_CONFIG'

ENV_MAP = {
    'MFS_CONFIG': MFS_CONFIG,
    'LSH_CONFIG': LSH_CONFIG,
    'EP_CONFIG': EP_CONFIG
}
queries_dir = f"queries_ep"
results_dir = f"results_ep"

env = load_environment(ENV_MAP[BU])
source_connection = env["source_connection"]
archived_connection = env["archived_connection"]
placeholders = env["placeholders"]

start_time = datetime.now()

wb = Workbook(write_only=True)
ws_summary = wb.create_sheet(title="Verification Results")
ws_summary.append([
    "Overall Result",
    "File",
    "Table/s",
    "Residual Count",
    "Residual Status",
    "Retained Count",
    "Expected Count",
    "Actual vs Expected Retained Result",
    "Data Integrity",
    "Timestamp"
])

sql_files = []
missing_files = []
for f in file_list:
    full_path = os.path.join(queries_dir, f)
    if not os.path.isfile(full_path):
        missing_files.append(f)
    else:
        sql_files.append(full_path)

if missing_files:
    print("\n❌ ERROR: The following files were not found in the 'queries' folder:")
    for mf in missing_files:
        print(f"   - {mf}")
    raise FileNotFoundError("One or more SQL files are missing. Please check the 'queries' folder.")

total_rows_verified = 0
total_columns_verified = 0
used_sheet_names = set()

def sanitize_sheet_name(name):
    valid_chars = "-_.() abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    return ''.join(c for c in name if c in valid_chars)

def get_unique_sheet_name(base_name):
    name = sanitize_sheet_name(base_name)[:31]
    original = name
    counter = 1
    while name in used_sheet_names:
        name = f"{original[:28]}_{counter}"[:31]
        counter += 1
    used_sheet_names.add(name)
    return name

for sql_file_path in sql_files:
    sql_file_name = os.path.basename(sql_file_path)

    try:
        with open(sql_file_path, 'r') as file:
            lines = file.readlines()
    except Exception as e:
        print(f"[X] Could not read file {sql_file_path}: {e}")
        continue

    print(f"\n📄 Verifying file: {sql_file_path}")
    sheet_title = get_unique_sheet_name(os.path.splitext(sql_file_name)[0])
    ws_data = wb.create_sheet(title=sheet_title)

    results_map = {}
    queries_data = {}
    column_names_retained = []
    column_names_expected = []
    i = 0

    while i < len(lines):
        if lines[i].strip().startswith("--"):
            label_line = lines[i].strip()[2:].strip()
            label = label_line.split('|')[0].strip().lower()

            i += 1
            table_names = ""
            if i < len(lines) and lines[i].strip().startswith("--"):
                table_names = lines[i].strip()[2:].strip()
                i += 1

            sql_lines = []
            while i < len(lines) and not lines[i].strip().startswith("--"):
                sql_lines.append(lines[i])
                i += 1
            sql_query = ''.join(sql_lines).strip()

            for placeholder, value in placeholders.items():
                sql_query = sql_query.replace(f"{{{placeholder}}}", value)

            if not sql_query:
                continue

            conn = archived_connection if 'retained' in label else source_connection

            try:
                with conn.cursor() as cursor:
                    cursor.execute(sql_query)
                    rows = cursor.fetchall()
                    count = len(rows)
                    print(f"[✓] {label.upper()} → {count}")

                    if (sql_file_name, table_names) not in results_map:
                        results_map[(sql_file_name, table_names)] = {
                            "residual_check": None,
                            "retained_count": None,
                            "expected_count": None
                        }

                    if 'residual' in label:
                        results_map[(sql_file_name, table_names)]["residual_check"] = count
                    elif 'retained' in label:
                        results_map[(sql_file_name, table_names)]["retained_count"] = count
                        queries_data['retained'] = list(rows)
                        column_names_retained = [desc[0] + "_retained" for desc in cursor.description]
                    elif 'expected' in label:
                        results_map[(sql_file_name, table_names)]["expected_count"] = count
                        queries_data['expected'] = list(rows)
                        column_names_expected = [desc[0] + "_expected" for desc in cursor.description]

            except Exception as e:
                print(f"[X] Error in {label.upper()} → {e}")
                ws_summary.append([
                    "FAIL",
                    sql_file_name,
                    table_names,
                    "Error" if 'residual' in label else "",
                    "FAIL",
                    "Error" if 'retained' in label else "",
                    "Error" if 'expected' in label else "",
                    "FAIL",
                    "Error: " + str(e),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
        else:
            i += 1

    for (sql_file, table), counts in results_map.items():
        residual = counts["residual_check"]
        retained = counts["retained_count"]
        expected = counts["expected_count"]

        residual_status = "PASS" if residual == 0 else "FAIL"

        retained_rows = queries_data.get("retained", [])
        expected_rows = queries_data.get("expected", [])

        mismatch_count = 0
        missing_count = 0

        interleaved_column_names = ["Overall Status"]
        for r_col, e_col in zip(column_names_retained, column_names_expected):
            col_base = r_col.replace("_retained", "")
            interleaved_column_names.append(f"{col_base}_status")
            interleaved_column_names.append(r_col)
            interleaved_column_names.append(e_col)
        ws_data.append(interleaved_column_names)

        base_col_names = [col.replace("_retained", "") for col in column_names_retained]
        key_indexes = [0] + [i for i, name in enumerate(base_col_names) if 'id' in name.lower()]

        def row_to_key(row, indexes):
            return tuple(row[i] for i in indexes)

        retained_dict = {row_to_key(row, key_indexes): row for row in retained_rows}
        expected_dict = {row_to_key(row, key_indexes): row for row in expected_rows}

        all_keys = set(retained_dict.keys()).union(expected_dict.keys())

        for key in all_keys:
            retained_row = retained_dict.get(key)
            expected_row = expected_dict.get(key)

            row_status = "PASS"
            interleaved_row = []

            if retained_row is None and expected_row is not None:
                row_status = "MISSING IN RETAINED"
                missing_count += 1
                retained_row = [""] * len(column_names_retained)
            elif expected_row is None and retained_row is not None:
                row_status = "MISSING IN EXPECTED"
                missing_count += 1
                expected_row = [""] * len(column_names_expected)

            interleaved_row.append(row_status)

            for r_val, e_val in zip(retained_row, expected_row):
                status = "matched" if r_val == e_val else "mismatched"
                if status == "mismatched" and row_status == "PASS":
                    row_status = "FAIL"
                    interleaved_row[0] = "FAIL"
                    mismatch_count += 1
                interleaved_row.append(status)
                interleaved_row.append(r_val)
                interleaved_row.append(e_val)

            total_rows_verified += 1
            cleaned_row = [clean_excel_value(cell) for cell in interleaved_row]
            ws_data.append(cleaned_row)

        if retained and expected:
            total_columns_verified += len(base_col_names)

        if (retained is None or retained == 0) and (expected is None or expected == 0):
            data_integrity_display = "No Data Fetched"
        elif mismatch_count == 0 and missing_count == 0 and retained == expected:
            data_integrity_display = "PASS"
        else:
            issues = []
            if mismatch_count > 0:
                issues.append(f"{mismatch_count} mismatches")
            if missing_count > 0:
                issues.append(f"{missing_count} missing rows")
            if retained != expected:
                issues.append(f"{abs(retained - expected)} row count diff")
            data_integrity_display = ", ".join(issues)

        retained_expected_status = "PASS" if mismatch_count == 0 and missing_count == 0 and retained == expected else "FAIL"

        if data_integrity_display == "No Data Fetched" and residual_status == "PASS" and retained_expected_status == "PASS":
            overall_result = "PASS"
        else:
            overall_result = "FAIL" if any([
                residual_status != "PASS",
                retained_expected_status != "PASS",
                data_integrity_display != "PASS"
            ]) else "PASS"

        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws_summary.append([
            overall_result,
            sql_file,
            table,
            residual if residual is not None else "",
            residual_status,
            retained if retained is not None else "",
            expected if expected is not None else "",
            retained_expected_status,
            data_integrity_display,
            timestamp_str
        ])

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
os.makedirs(results_dir, exist_ok=True)
output_file = os.path.join(results_dir, f"{BU}_query_results_{timestamp}.xlsx")
wb.save(output_file)

duration = datetime.now() - start_time
total_seconds = duration.total_seconds()
milliseconds = int((total_seconds - int(total_seconds)) * 1000)
formatted_duration = f"{str(duration).split('.')[0]}.{milliseconds:03d}"

print(f"\n📉 Results saved to: {output_file}")
print(f"⏱️ Duration: {formatted_duration}")
print(f"🔢 Total rows verified: {total_rows_verified}")
print(f"📊 Total columns verified: {total_columns_verified}")
